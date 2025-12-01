"""Excel reader implementation for .xlsx files using openpyxl."""

from typing import List, Dict, Any

from src.excel_tools.base import ExcelReaderBase, CellData, CellRange, Direction


class OpenpyxlReader(ExcelReaderBase):
    """Excel reader implementation for .xlsx files using openpyxl."""

    def __init__(self, file_path: str):
        """Initialize the openpyxl reader."""
        super().__init__(file_path)
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Fill, Border
        except ImportError:
            raise ImportError(
                "openpyxl is required for .xlsx files. Install it with: pip install openpyxl"
            )

        self.workbook = load_workbook(file_path, data_only=True)

    def get_sheet_names(self) -> List[str]:
        """Get list of all sheet names in the workbook."""
        return self.workbook.sheetnames

    def get_sheet_bounds(self, sheet_name: str) -> str:
        """Get the used range of a sheet in Excel notation."""
        sheet = self.workbook[sheet_name]

        if sheet.max_row == 0 or sheet.max_column == 0:
            return "A1:A1"

        # Get the actual used range
        min_row = sheet.min_row
        min_col = sheet.min_column
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Convert to Excel notation
        start_cell = self._get_column_letter(min_col) + str(min_row)
        end_cell = self._get_column_letter(max_col) + str(max_row)

        return f"{start_cell}:{end_cell}"

    def get_cells_in_range(self, sheet_name: str, range_notation: str) -> List[CellData]:
        """Get cells with values and formatting in the specified range."""
        sheet = self.workbook[sheet_name]
        cell_range = CellRange.from_excel_notation(range_notation)

        cells = []
        for row in range(cell_range.start_row, cell_range.end_row + 1):
            for col in range(cell_range.start_col, cell_range.end_col + 1):
                cell = sheet.cell(row + 1, col + 1)  # openpyxl uses 1-indexed
                address = self._get_column_letter(col + 1) + str(row + 1)

                # Extract formatting information
                formatting = self._get_cell_formatting(cell)

                cells.append(CellData(address=address, value=cell.value, formatting=formatting))

        return cells

    def iterate_until_empty(
        self, sheet_name: str, start_cell: str, direction: Direction
    ) -> List[CellData]:
        """Iterate from a cell in a direction until an empty cell is found."""
        sheet = self.workbook[sheet_name]
        col, row = CellRange._parse_cell_address(start_cell)

        cells = []

        # Define direction deltas
        deltas = {
            Direction.UP: (0, -1),
            Direction.DOWN: (0, 1),
            Direction.LEFT: (-1, 0),
            Direction.RIGHT: (1, 0),
        }

        delta_col, delta_row = deltas[direction]

        while True:
            # Get the cell at current position
            cell = sheet.cell(row + 1, col + 1)  # openpyxl uses 1-indexed

            # Check if cell is empty
            if cell.value is None or str(cell.value).strip() == "":
                break

            # Add cell to results
            address = self._get_column_letter(col + 1) + str(row + 1)
            formatting = self._get_cell_formatting(cell)
            cells.append(CellData(address=address, value=cell.value, formatting=formatting))

            # Move to next cell
            col += delta_col
            row += delta_row

            # Check bounds (col/row are 0-indexed, max_column/max_row are 1-indexed)
            if col < 0 or row < 0:
                break
            if col + 1 > sheet.max_column or row + 1 > sheet.max_row:
                break

        return cells

    def close(self):
        """Close the workbook and free resources."""
        self.workbook.close()

    @staticmethod
    def _get_column_letter(col: int) -> str:
        """Convert column number to Excel letter (1 -> A, 2 -> B, etc.)."""
        from openpyxl.utils import get_column_letter

        return get_column_letter(col)

    @staticmethod
    def _get_cell_formatting(cell) -> Dict[str, Any]:
        """Extract formatting information from a cell."""
        formatting = {}

        # Font information
        if cell.font:
            formatting["bold"] = cell.font.bold
            formatting["italic"] = cell.font.italic
            formatting["underline"] = cell.font.underline is not None
            formatting["font_size"] = cell.font.size
            formatting["font_color"] = (
                str(cell.font.color.rgb)
                if cell.font.color and hasattr(cell.font.color, "rgb")
                else None
            )

        # Fill (background) color
        if cell.fill:
            formatting["fill_color"] = (
                str(cell.fill.fgColor.rgb)
                if cell.fill.fgColor and hasattr(cell.fill.fgColor, "rgb")
                else None
            )

        # Border information
        if cell.border:
            formatting["has_border"] = any(
                [
                    cell.border.left and cell.border.left.style,
                    cell.border.right and cell.border.right.style,
                    cell.border.top and cell.border.top.style,
                    cell.border.bottom and cell.border.bottom.style,
                ]
            )

        # Number format
        formatting["number_format"] = cell.number_format

        # Alignment
        if cell.alignment:
            formatting["horizontal_alignment"] = cell.alignment.horizontal
            formatting["vertical_alignment"] = cell.alignment.vertical

        return formatting
