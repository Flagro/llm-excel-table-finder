"""Abstract base class and tools for Excel operations."""

from abc import ABC, abstractmethod
from typing import List, Dict, Any, Tuple, Optional
from dataclasses import dataclass
from enum import Enum


class Direction(str, Enum):
    """Direction for cell iteration."""

    UP = "up"
    DOWN = "down"
    LEFT = "left"
    RIGHT = "right"


@dataclass
class CellData:
    """Data structure for cell information."""

    address: str
    value: Any
    formatting: Dict[str, Any]


@dataclass
class CellRange:
    """Data structure for cell range."""

    start_col: int  # 0-indexed
    start_row: int  # 0-indexed
    end_col: int  # 0-indexed
    end_row: int  # 0-indexed

    def to_excel_notation(self) -> str:
        """Convert to Excel A1 notation (e.g., A3:C10)."""
        start = self._to_column_letter(self.start_col) + str(self.start_row + 1)
        end = self._to_column_letter(self.end_col) + str(self.end_row + 1)
        return f"{start}:{end}"

    @staticmethod
    def _to_column_letter(col: int) -> str:
        """Convert column index to Excel letter (0 -> A, 1 -> B, etc.)."""
        result = ""
        while col >= 0:
            result = chr(col % 26 + ord("A")) + result
            col = col // 26 - 1
        return result

    @classmethod
    def from_excel_notation(cls, notation: str) -> "CellRange":
        """Parse Excel A1 notation (e.g., A3:C10) into CellRange."""
        if ":" not in notation:
            # Single cell
            col, row = cls._parse_cell_address(notation)
            return cls(col, row, col, row)

        start_cell, end_cell = notation.split(":")
        start_col, start_row = cls._parse_cell_address(start_cell)
        end_col, end_row = cls._parse_cell_address(end_cell)

        return cls(start_col, start_row, end_col, end_row)

    @staticmethod
    def _parse_cell_address(address: str) -> Tuple[int, int]:
        """Parse cell address like 'A3' into (col_idx, row_idx)."""
        col_letters = ""
        row_digits = ""

        for char in address:
            if char.isalpha():
                col_letters += char.upper()
            elif char.isdigit():
                row_digits += char

        # Convert column letters to index
        col_idx = 0
        for char in col_letters:
            col_idx = col_idx * 26 + (ord(char) - ord("A") + 1)
        col_idx -= 1  # Convert to 0-indexed

        row_idx = int(row_digits) - 1  # Convert to 0-indexed

        return col_idx, row_idx


class ExcelReaderBase(ABC):
    """Abstract base class for Excel file operations."""

    def __init__(self, file_path: str):
        """Initialize the Excel reader with a file path."""
        self.file_path = file_path

    @abstractmethod
    def get_sheet_names(self) -> List[str]:
        """Get list of all sheet names in the workbook."""
        pass

    @abstractmethod
    def get_sheet_bounds(self, sheet_name: str) -> str:
        """
        Get the used range of a sheet in Excel notation.

        Args:
            sheet_name: Name of the sheet

        Returns:
            Range in Excel notation (e.g., "A1:Z100")
        """
        pass

    @abstractmethod
    def get_cells_in_range(self, sheet_name: str, range_notation: str) -> List[CellData]:
        """
        Get cells with values and formatting in the specified range.

        Args:
            sheet_name: Name of the sheet
            range_notation: Range in Excel notation (e.g., "A3:C10")

        Returns:
            List of CellData objects
        """
        pass

    @abstractmethod
    def iterate_until_empty(
        self, sheet_name: str, start_cell: str, direction: Direction
    ) -> List[CellData]:
        """
        Iterate from a cell in a direction until an empty cell is found.

        Args:
            sheet_name: Name of the sheet
            start_cell: Starting cell in Excel notation (e.g., "A3")
            direction: Direction to iterate (up, down, left, right)

        Returns:
            List of CellData objects encountered (excluding the empty cell)
        """
        pass

    @abstractmethod
    def close(self):
        """Close the workbook and free resources."""
        pass

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        self.close()


class OpenpyxlReader(ExcelReaderBase):
    """Excel reader implementation for .xlsx files using openpyxl."""

    def __init__(self, file_path: str):
        """Initialize the openpyxl reader."""
        super().__init__(file_path)
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Fill, Border
        except ImportError:
            raise ImportError(
                "openpyxl is required for .xlsx files. Install it with: pip install openpyxl"
            )

        self.workbook = load_workbook(file_path, data_only=True)

    def get_sheet_names(self) -> List[str]:
        """Get list of all sheet names in the workbook."""
        return self.workbook.sheetnames

    def get_sheet_bounds(self, sheet_name: str) -> str:
        """Get the used range of a sheet in Excel notation."""
        sheet = self.workbook[sheet_name]

        if sheet.max_row == 0 or sheet.max_column == 0:
            return "A1:A1"

        # Get the actual used range
        min_row = sheet.min_row
        min_col = sheet.min_column
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Convert to Excel notation
        start_cell = self._get_column_letter(min_col) + str(min_row)
        end_cell = self._get_column_letter(max_col) + str(max_row)

        return f"{start_cell}:{end_cell}"

    def get_cells_in_range(self, sheet_name: str, range_notation: str) -> List[CellData]:
        """Get cells with values and formatting in the specified range."""
        sheet = self.workbook[sheet_name]
        cell_range = CellRange.from_excel_notation(range_notation)

        cells = []
        for row in range(cell_range.start_row, cell_range.end_row + 1):
            for col in range(cell_range.start_col, cell_range.end_col + 1):
                cell = sheet.cell(row + 1, col + 1)  # openpyxl uses 1-indexed
                address = self._get_column_letter(col + 1) + str(row + 1)

                # Extract formatting information
                formatting = self._get_cell_formatting(cell)

                cells.append(CellData(address=address, value=cell.value, formatting=formatting))

        return cells

    def iterate_until_empty(
        self, sheet_name: str, start_cell: str, direction: Direction
    ) -> List[CellData]:
        """Iterate from a cell in a direction until an empty cell is found."""
        sheet = self.workbook[sheet_name]
        col, row = CellRange._parse_cell_address(start_cell)

        cells = []

        # Define direction deltas
        deltas = {
            Direction.UP: (0, -1),
            Direction.DOWN: (0, 1),
            Direction.LEFT: (-1, 0),
            Direction.RIGHT: (1, 0),
        }

        delta_col, delta_row = deltas[direction]

        while True:
            # Get the cell at current position
            cell = sheet.cell(row + 1, col + 1)  # openpyxl uses 1-indexed

            # Check if cell is empty
            if cell.value is None or str(cell.value).strip() == "":
                break

            # Add cell to results
            address = self._get_column_letter(col + 1) + str(row + 1)
            formatting = self._get_cell_formatting(cell)
            cells.append(CellData(address=address, value=cell.value, formatting=formatting))

            # Move to next cell
            col += delta_col
            row += delta_row

            # Check bounds
            if col < 0 or row < 0:
                break
            if col >= sheet.max_column or row >= sheet.max_row:
                break

        return cells

    def close(self):
        """Close the workbook and free resources."""
        self.workbook.close()

    @staticmethod
    def _get_column_letter(col: int) -> str:
        """Convert column number to Excel letter (1 -> A, 2 -> B, etc.)."""
        from openpyxl.utils import get_column_letter

        return get_column_letter(col)

    @staticmethod
    def _get_cell_formatting(cell) -> Dict[str, Any]:
        """Extract formatting information from a cell."""
        formatting = {}

        # Font information
        if cell.font:
            formatting["bold"] = cell.font.bold
            formatting["italic"] = cell.font.italic
            formatting["underline"] = cell.font.underline is not None
            formatting["font_size"] = cell.font.size
            formatting["font_color"] = (
                str(cell.font.color.rgb)
                if cell.font.color and hasattr(cell.font.color, "rgb")
                else None
            )

        # Fill (background) color
        if cell.fill:
            formatting["fill_color"] = (
                str(cell.fill.fgColor.rgb)
                if cell.fill.fgColor and hasattr(cell.fill.fgColor, "rgb")
                else None
            )

        # Border information
        if cell.border:
            formatting["has_border"] = any(
                [
                    cell.border.left and cell.border.left.style,
                    cell.border.right and cell.border.right.style,
                    cell.border.top and cell.border.top.style,
                    cell.border.bottom and cell.border.bottom.style,
                ]
            )

        # Number format
        formatting["number_format"] = cell.number_format

        # Alignment
        if cell.alignment:
            formatting["horizontal_alignment"] = cell.alignment.horizontal
            formatting["vertical_alignment"] = cell.alignment.vertical

        return formatting


class XlrdReader(ExcelReaderBase):
    """Excel reader implementation for .xls files using xlrd."""

    def __init__(self, file_path: str):
        """Initialize the xlrd reader."""
        super().__init__(file_path)
        try:
            import xlrd
            from xlrd.formatting import Format, XF
        except ImportError:
            raise ImportError("xlrd is required for .xls files. Install it with: pip install xlrd")

        self.workbook = xlrd.open_workbook(file_path, formatting_info=True)

    def get_sheet_names(self) -> List[str]:
        """Get list of all sheet names in the workbook."""
        return self.workbook.sheet_names()

    def get_sheet_bounds(self, sheet_name: str) -> str:
        """Get the used range of a sheet in Excel notation."""
        sheet = self.workbook.sheet_by_name(sheet_name)

        if sheet.nrows == 0 or sheet.ncols == 0:
            return "A1:A1"

        # xlrd provides nrows and ncols (0-indexed counts)
        max_row = sheet.nrows
        max_col = sheet.ncols

        # Convert to Excel notation
        start_cell = "A1"
        end_cell = CellRange._to_column_letter(max_col - 1) + str(max_row)

        return f"{start_cell}:{end_cell}"

    def get_cells_in_range(self, sheet_name: str, range_notation: str) -> List[CellData]:
        """Get cells with values and formatting in the specified range."""
        sheet = self.workbook.sheet_by_name(sheet_name)
        cell_range = CellRange.from_excel_notation(range_notation)

        cells = []
        for row in range(cell_range.start_row, min(cell_range.end_row + 1, sheet.nrows)):
            for col in range(cell_range.start_col, min(cell_range.end_col + 1, sheet.ncols)):
                cell = sheet.cell(row, col)
                address = CellRange._to_column_letter(col) + str(row + 1)

                # Extract formatting information
                formatting = self._get_cell_formatting(sheet, row, col)

                # Get cell value
                value = self._get_cell_value(cell)

                cells.append(CellData(address=address, value=value, formatting=formatting))

        return cells

    def iterate_until_empty(
        self, sheet_name: str, start_cell: str, direction: Direction
    ) -> List[CellData]:
        """Iterate from a cell in a direction until an empty cell is found."""
        sheet = self.workbook.sheet_by_name(sheet_name)
        col, row = CellRange._parse_cell_address(start_cell)

        cells = []

        # Define direction deltas
        deltas = {
            Direction.UP: (0, -1),
            Direction.DOWN: (0, 1),
            Direction.LEFT: (-1, 0),
            Direction.RIGHT: (1, 0),
        }

        delta_col, delta_row = deltas[direction]

        while True:
            # Check bounds
            if row < 0 or col < 0 or row >= sheet.nrows or col >= sheet.ncols:
                break

            # Get the cell at current position
            cell = sheet.cell(row, col)
            value = self._get_cell_value(cell)

            # Check if cell is empty
            if value is None or str(value).strip() == "":
                break

            # Add cell to results
            address = CellRange._to_column_letter(col) + str(row + 1)
            formatting = self._get_cell_formatting(sheet, row, col)
            cells.append(CellData(address=address, value=value, formatting=formatting))

            # Move to next cell
            col += delta_col
            row += delta_row

        return cells

    def close(self):
        """Close the workbook and free resources."""
        # xlrd doesn't require explicit closing, but we set to None for consistency
        self.workbook = None

    @staticmethod
    def _get_cell_value(cell):
        """Get the actual value from an xlrd cell."""
        import xlrd

        # Handle different cell types
        if cell.ctype == xlrd.XL_CELL_EMPTY:
            return None
        elif cell.ctype == xlrd.XL_CELL_TEXT:
            return cell.value
        elif cell.ctype == xlrd.XL_CELL_NUMBER:
            return cell.value
        elif cell.ctype == xlrd.XL_CELL_DATE:
            # Convert Excel date to a tuple
            date_tuple = xlrd.xldate_as_tuple(cell.value, 0)
            return f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(cell.value)
        else:
            return cell.value

    def _get_cell_formatting(self, sheet, row: int, col: int) -> Dict[str, Any]:
        """Extract formatting information from a cell."""
        import xlrd

        formatting = {}

        try:
            # Get the cell's XF record
            cell = sheet.cell(row, col)
            xf_index = cell.xf_index

            if xf_index is not None:
                xf = self.workbook.format_map.get(self.workbook.xf_list[xf_index].format_key)

                # Get font information
                font = self.workbook.font_list[self.workbook.xf_list[xf_index].font_index]
                formatting["bold"] = bool(font.bold)
                formatting["italic"] = bool(font.italic)
                formatting["underline"] = bool(font.underline_type)
                formatting["font_size"] = font.height / 20  # Convert to points

                # Get number format
                if xf:
                    formatting["number_format"] = xf.format_str

                # Get alignment
                alignment = self.workbook.xf_list[xf_index].alignment
                formatting["horizontal_alignment"] = alignment.hor_align
                formatting["vertical_alignment"] = alignment.vert_align

                # Background color/pattern
                background = self.workbook.xf_list[xf_index].background
                formatting["background_color_index"] = background.pattern_colour_index

        except (AttributeError, IndexError, KeyError):
            # If we can't get formatting info, just return empty dict
            pass

        return formatting
